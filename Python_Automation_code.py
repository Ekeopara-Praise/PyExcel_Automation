# import the necessary library
import pandas as pd
from openpyxl import load_workbook

# define number of experiments 
num_exp = 2
count = 1 

while count <= num_exp:   # create the while loop statement
    
    # load excel file
    workbook_source = load_workbook(filename=r"C:\Users\Sir_Praise\Documents\Datasets\Cyril_file.xlsx") # source excel file 
    workbook_data = load_workbook(filename=r"C:\Users\Sir_Praise\Documents\Datasets\Cyril_simulationdata.xlsx")  # data excel file


    # open workbooks
    sheet_source = workbook_source.active
    sheet_data = workbook_data.active
    
    i = count + 2 # define rows in the workbook
    
    # modify the desired cell
   
    sheet_source["A405"] = f"    'PROD1' 'G'     {sheet_data[f'B{i}'].value}  {sheet_data[f'C{i}'].value}   8400 'OIL'  /"
    sheet_source["A406"] = f"    'PROD2' 'G'    {sheet_data[f'D{i}'].value}  {sheet_data[f'E{i}'].value}   8400 'OIL'  /" 
    sheet_source["A407"] = f"    'PROD3' 'G'    {sheet_data[f'F{i}'].value} {sheet_data[f'G{i}'].value}   8400 'OIL'  /"
    sheet_source["A408"] = f"    'PROD4' 'G'     {sheet_data[f'H{i}'].value} {sheet_data[f'I{i}'].value}   8400 'OIL'  /" 
    sheet_source["A409"] = f"    'INJ1'  'G'    {sheet_data[f'J{i}'].value} {sheet_data[f'K{i}'].value}   8400 'WAT'  /"


    sheet_source["A417"] = f"    'PROD1'   {sheet_data[f'B{i}'].value}  {sheet_data[f'C{i}'].value}  1  1 'OPEN' 0   -1   0.5  /"
    sheet_source["A418"] = f"    'PROD2'  {sheet_data[f'D{i}'].value}  {sheet_data[f'E{i}'].value}  1  1 'OPEN' 0   -1   0.5  /"
    sheet_source["A419"] = f"    'PROD3'  {sheet_data[f'F{i}'].value} {sheet_data[f'G{i}'].value}  1  1 'OPEN' 0   -1   0.5  /"
    sheet_source["A420"] = f"    'PROD4'   {sheet_data[f'H{i}'].value} {sheet_data[f'I{i}'].value}  1  1 'OPEN' 0   -1   0.5  /"
    sheet_source["A421"] = f"    'INJ1'   {sheet_data[f'J{i}'].value} {sheet_data[f'K{i}'].value}  1  1 'OPEN' 1   -1   0.5  /"



    # define filename
    filename = rf"C:\Users\Sir_Praise\Documents\Datasets\Saved data\RUN_{count}.xlsx"

    # save the file
    workbook_source.save(filename=filename)
    
    # read the data into a pandas dataframe and save to .DATA extension
    file = pd.read_excel(filename)
    file.to_csv(rf'C:\Users\Sir_Praise\Documents\Datasets\Data Files\RUN_{count}'+'.DATA', index=False)
    
    count = count + 1  # incrementing the count after each successful counts